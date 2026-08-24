"""
NSE Streak Report — FastAPI backend
-------------------------------------
Serves a minimal web UI and a JSON API that detects N-day consecutive
up/down streaks ending on a given date, and reports the next trading
day's move as a "verdict" (continued / broken / paused).

Fetches price history live from Yahoo Finance on each request. No local
caching/database -- purely a web app.

Run:
    pip install -r requirements.txt
    uvicorn main:app --reload
Then open http://127.0.0.1:8000
"""

from datetime import datetime, timedelta
from io import BytesIO
from typing import List, Optional, Literal
from concurrent.futures import ThreadPoolExecutor, as_completed
import csv
import io
import json

import pandas as pd
import requests
import yfinance as yf
from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

app = FastAPI(title="NSE Streak Report")

# ----------------------------------------------------------------------
# NSE / Yahoo Finance data fetching
# ----------------------------------------------------------------------

NSE_EQUITY_LIST_URL = "https://nsearchives.nseindia.com/content/equities/EQUITY_L.csv"

# Small fallback list, used only if the live NSE list can't be fetched
# (e.g. no internet, or NSE temporarily blocks the request).
FALLBACK_SYMBOLS = [
    "RELIANCE", "TCS", "INFY", "HDFCBANK", "ICICIBANK", "HINDUNILVR", "SBIN",
    "BHARTIARTL", "ITC", "KOTAKBANK", "LT", "AXISBANK", "BAJFINANCE", "ASIANPAINT",
    "MARUTI", "SUNPHARMA", "TITAN", "ULTRACEMCO", "WIPRO", "ONGC", "NESTLEIND",
    "HCLTECH", "BAJAJFINSV", "ADANIENT", "ADANIPORTS", "POWERGRID", "NTPC",
    "TATASTEEL", "TATAMOTORS", "JSWSTEEL", "M&M", "COALINDIA", "GRASIM", "TECHM",
    "HDFCLIFE", "SBILIFE", "DIVISLAB", "DRREDDY", "CIPLA", "EICHERMOT", "BRITANNIA",
    "APOLLOHOSP", "HEROMOTOCO", "BAJAJ-AUTO", "INDUSINDBK", "SHRIRAMFIN",
    "TATACONSUM", "UPL", "BPCL", "LTIM", "PIDILITIND", "DABUR", "GODREJCP",
    "MARICO", "COLPAL", "HAVELLS", "SIEMENS", "ABB", "DLF", "GAIL", "IOC",
    "HINDPETRO", "VEDL", "HINDALCO", "JINDALSTEL", "SAIL", "BANKBARODA", "PNB",
    "CANBK", "IDFCFIRSTB", "FEDERALBNK", "AUBANK", "BANDHANBNK", "PEL",
    "MUTHOOTFIN", "CHOLAFIN", "LICHSGFIN", "RECLTD", "PFC", "IRFC", "ZOMATO",
    "NYKAA", "PAYTM", "POLICYBZR", "IRCTC", "DMART", "TRENT", "NAUKRI",
    "PVRINOX", "INDIGO", "AMBUJACEM", "ACC", "SHREECEM", "RAMCOCEM", "DALBHARAT",
    "MOTHERSON", "BOSCHLTD", "MRF", "BALKRISIND", "EXIDEIND",
]
FALLBACK_ENTRIES = [{"symbol": s, "listing_date": None} for s in FALLBACK_SYMBOLS]

# Display name -> CSV filename for preset index constituent lists.
# These follow NSE's standard naming pattern (confirmed working for
# ind_nifty50list.csv); a few of the less common sector ones are
# best-effort and may need a filename correction if NSE has renamed them.
INDEX_LIST_BASE = "https://nsearchives.nseindia.com/content/indices/"
INDEX_PRESETS = {
    "Nifty 50": "ind_nifty50list.csv",
    "Nifty Next 50": "ind_niftynext50list.csv",
    "Nifty 100": "ind_nifty100list.csv",
    "Nifty 200": "ind_nifty200list.csv",
    "Nifty 500": "ind_nifty500list.csv",
    "Nifty Bank": "ind_niftybanklist.csv",
    "Nifty IT": "ind_niftyitlist.csv",
    "Nifty Auto": "ind_niftyautolist.csv",
    "Nifty Pharma": "ind_niftypharmalist.csv",
    "Nifty FMCG": "ind_niftyfmcglist.csv",
    "Nifty Infra": "ind_niftyinfralist.csv",
    "Nifty Metal": "ind_niftymetallist.csv",
    "Nifty Energy": "ind_niftyenergylist.csv",
    "Nifty Realty": "ind_niftyrealtylist.csv",
}

_NSE_REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def _nse_session() -> requests.Session:
    """NSE's site blocks bare requests, so we warm up a session against
    the homepage first (like a browser would) to pick up the cookies it
    expects before hitting the actual data endpoint."""
    session = requests.Session()
    session.headers.update(_NSE_REQUEST_HEADERS)
    session.get("https://www.nseindia.com/", timeout=10)
    return session


def fetch_all_nse_entries() -> List[dict]:
    """Fetch NSE's official list of all listed equities (symbol + listing
    date). Falls back to a curated ~100-stock list if the request fails."""
    try:
        session = _nse_session()
        resp = session.get(NSE_EQUITY_LIST_URL, timeout=20)
        resp.raise_for_status()
        reader = csv.DictReader(io.StringIO(resp.text))
        fieldnames = reader.fieldnames or []
        symbol_key = next((f for f in fieldnames if f.strip().upper() == "SYMBOL"), "SYMBOL")
        listing_key = next((f for f in fieldnames if f.strip().upper() == "DATE OF LISTING"), None)

        entries = []
        for row in reader:
            symbol = (row.get(symbol_key) or "").strip()
            if not symbol:
                continue
            listing_date = None
            raw_date = (row.get(listing_key) or "").strip() if listing_key else ""
            if raw_date:
                try:
                    listing_date = pd.Timestamp(datetime.strptime(raw_date, "%d-%b-%Y"))
                except ValueError:
                    listing_date = None
            entries.append({"symbol": symbol, "listing_date": listing_date})

        if not entries:
            raise ValueError("Parsed symbol list was empty")
        return entries
    except Exception as e:
        print(f"Could not fetch full NSE symbol list ({e}); using fallback list.")
        return FALLBACK_ENTRIES


def fetch_symbol_history(symbol: str, start: str, end: str) -> pd.DataFrame:
    """Fetch daily OHLCV history for symbol (no .NS suffix) between
    start and end (YYYY-MM-DD, end exclusive per yfinance convention).
    Returns a DataFrame indexed by tz-naive date, or empty on failure."""
    try:
        hist = yf.Ticker(f"{symbol}.NS").history(start=start, end=end)
        if hist.empty:
            return hist
        if hist.index.tz is not None:
            hist.index = hist.index.tz_localize(None)
        return hist
    except Exception:
        return pd.DataFrame()


def fetch_index_constituents(index_name: str) -> List[str]:
    """Fetch the current constituent symbols for a named preset index
    (e.g. 'Nifty 50', 'Nifty Bank'). Raises ValueError if the name isn't
    a known preset, or RuntimeError if the fetch/parse fails."""
    filename = INDEX_PRESETS.get(index_name)
    if not filename:
        raise ValueError(f"Unknown index preset: {index_name}")

    try:
        session = _nse_session()
        resp = session.get(INDEX_LIST_BASE + filename, timeout=20)
        resp.raise_for_status()
        reader = csv.DictReader(io.StringIO(resp.text))
        fieldnames = reader.fieldnames or []
        symbol_key = next((f for f in fieldnames if f.strip().upper() == "SYMBOL"), None)
        if not symbol_key:
            raise RuntimeError(f"Unexpected CSV format for {index_name}")

        symbols = [row[symbol_key].strip() for row in reader if row.get(symbol_key, "").strip()]
        if not symbols:
            raise RuntimeError(f"No symbols parsed for {index_name}")
        return symbols
    except Exception as e:
        raise RuntimeError(f"Could not fetch {index_name} constituents: {e}")


_symbol_cache = {"entries": None, "fetched_at": None}


def get_all_nse_entries() -> List[dict]:
    """Live NSE entries (symbol + listing date), cached in-process for 24h."""
    now = datetime.utcnow()
    cached_at = _symbol_cache["fetched_at"]
    if _symbol_cache["entries"] is None or cached_at is None or (now - cached_at).total_seconds() > 86400:
        _symbol_cache["entries"] = fetch_all_nse_entries()
        _symbol_cache["fetched_at"] = now
    return _symbol_cache["entries"]


def get_default_symbols(end_date: Optional[str] = None) -> List[str]:
    """Returns NSE symbols, automatically excluding stocks that weren't
    listed yet as of end_date (if given) -- so a report for e.g. 2015
    won't waste time checking companies that IPO'd years later."""
    entries = get_all_nse_entries()
    if not end_date:
        return [e["symbol"] for e in entries]

    cutoff = pd.Timestamp(end_date)
    return [
        e["symbol"] for e in entries
        if e["listing_date"] is None or e["listing_date"] <= cutoff
    ]


# ----------------------------------------------------------------------
# Request / response models
# ----------------------------------------------------------------------

class ReportRequest(BaseModel):
    end_date: str = Field(..., description="YYYY-MM-DD, last day of the streak window")
    n_days: int = Field(..., ge=1, le=30)
    symbols: Optional[List[str]] = None  # if omitted, uses DEFAULT_SYMBOLS


class DayChange(BaseModel):
    label: str
    pct_change: float
    volume: int
    volume_pct_change: Optional[float] = None


class StockResult(BaseModel):
    symbol: str
    direction: Literal["Up", "Down"]
    net_pct_change: float
    days: List[DayChange]
    next_day_label: Optional[str]
    next_day_pct_change: Optional[float]
    next_day_volume: Optional[int]
    verdict: str


class ReportResponse(BaseModel):
    up_streaks: List[StockResult]
    down_streaks: List[StockResult]
    checked: int
    skipped: int


# ----------------------------------------------------------------------
# Core streak logic
# ----------------------------------------------------------------------

def compute_streak(closes: List[float]):
    directions = []
    for i in range(1, len(closes)):
        if closes[i] > closes[i - 1]:
            directions.append("Up")
        elif closes[i] < closes[i - 1]:
            directions.append("Down")
        else:
            directions.append("Flat")

    if not directions:
        return 0, "Flat"

    last_dir = directions[-1]
    if last_dir == "Flat":
        return 0, "Flat"

    streak = 0
    for d in reversed(directions):
        if d == last_dir:
            streak += 1
        else:
            break
    return streak, last_dir


def evaluate_symbol(symbol: str, end_date_str: str, n_days: int) -> Optional[StockResult]:
    end_dt = pd.Timestamp(end_date_str)
    fetch_start = (end_dt - timedelta(days=n_days * 3 + 25)).strftime("%Y-%m-%d")
    fetch_end = (end_dt + timedelta(days=15)).strftime("%Y-%m-%d")

    hist = fetch_symbol_history(symbol, fetch_start, fetch_end)
    if hist.empty:
        return None

    valid_dates = hist.index[hist.index <= end_dt]
    if valid_dates.empty:
        return None
    end_pos = hist.index.get_loc(valid_dates[-1])
    if end_pos < n_days:
        return None

    window = hist.iloc[end_pos - n_days: end_pos + 1]
    window_rows = [
        {"date": idx.strftime("%Y-%m-%d"), "close": float(row["Close"]), "volume": int(row["Volume"])}
        for idx, row in window.iterrows()
    ]

    next_row = None
    if end_pos + 1 < len(hist):
        next_idx = hist.index[end_pos + 1]
        next_series = hist.iloc[end_pos + 1]
        next_row = {
            "date": next_idx.strftime("%Y-%m-%d"),
            "close": float(next_series["Close"]),
            "volume": int(next_series["Volume"]),
        }

    return _build_result_from_rows(symbol, window_rows, next_row)


def _build_result_from_rows(symbol: str, window_rows: List[dict], next_row: Optional[dict]) -> Optional[StockResult]:
    """Given a resolved n_days+1 window of rows (oldest first) plus an
    optional next-trading-day row, compute the streak and build a
    StockResult."""
    closes = [r["close"] for r in window_rows]
    streak_len, direction = compute_streak(closes)

    if streak_len < len(window_rows) - 1 or direction == "Flat":
        return None

    net_pct = (closes[-1] - closes[0]) / closes[0] * 100

    days = []
    for i in range(1, len(window_rows)):
        prev_close = closes[i - 1]
        this_close = closes[i]
        day_pct = (this_close - prev_close) / prev_close * 100

        prev_volume = window_rows[i - 1]["volume"]
        this_volume = window_rows[i]["volume"]
        volume_pct = ((this_volume - prev_volume) / prev_volume * 100) if prev_volume else None

        row = window_rows[i]
        date_label = row["date"] if isinstance(row["date"], str) else row["date"].strftime("%Y-%m-%d")
        days.append(DayChange(
            label=f"Day {i} ({_pretty_date(date_label)})",
            pct_change=round(day_pct, 2),
            volume=int(row["volume"]),
            volume_pct_change=round(volume_pct, 2) if volume_pct is not None else None,
        ))

    next_day_label = None
    next_day_pct = None
    next_day_vol = None
    verdict = "No data yet for next trading day"

    if next_row is not None:
        last_close = closes[-1]
        next_pct = (next_row["close"] - last_close) / last_close * 100
        next_dir = "Up" if next_pct > 0 else ("Down" if next_pct < 0 else "Flat")

        if next_dir == direction:
            verdict = f"Streak continued ({direction})"
        elif next_dir == "Flat":
            verdict = "Streak paused (flat)"
        else:
            verdict = f"Streak broken (reversed to {next_dir})"

        next_date_label = next_row["date"] if isinstance(next_row["date"], str) else next_row["date"].strftime("%Y-%m-%d")
        next_day_label = f"Next day ({_pretty_date(next_date_label)})"
        next_day_pct = round(next_pct, 2)
        next_day_vol = int(next_row["volume"])

    return StockResult(
        symbol=symbol,
        direction=direction,
        net_pct_change=round(net_pct, 2),
        days=days,
        next_day_label=next_day_label,
        next_day_pct_change=next_day_pct,
        next_day_volume=next_day_vol,
        verdict=verdict,
    )


def _pretty_date(date_str: str) -> str:
    return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d-%b-%Y")


# ----------------------------------------------------------------------
# API routes
# ----------------------------------------------------------------------

def build_report(req: "ReportRequest") -> ReportResponse:
    try:
        datetime.strptime(req.end_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="end_date must be in YYYY-MM-DD format")

    symbols = req.symbols if req.symbols else get_default_symbols(req.end_date)
    symbols = [s.strip().upper() for s in symbols if s.strip()]

    up_streaks, down_streaks = [], []
    skipped = 0

    # Parallelize: the full NSE list runs into the thousands of symbols,
    # and each is a separate network call, so fetch concurrently.
    with ThreadPoolExecutor(max_workers=20) as pool:
        futures = {
            pool.submit(evaluate_symbol, sym, req.end_date, req.n_days): sym
            for sym in symbols
        }
        for future in as_completed(futures):
            try:
                result = future.result()
            except Exception:
                result = None
            if result is None:
                skipped += 1
                continue
            if result.direction == "Up":
                up_streaks.append(result)
            else:
                down_streaks.append(result)

    up_streaks.sort(key=lambda r: r.net_pct_change, reverse=True)
    down_streaks.sort(key=lambda r: r.net_pct_change)

    return ReportResponse(
        up_streaks=up_streaks,
        down_streaks=down_streaks,
        checked=len(symbols),
        skipped=skipped,
    )


# ----------------------------------------------------------------------
# Excel export
# ----------------------------------------------------------------------

def _write_sheet(wb: Workbook, sheet_name: str, stocks: List[StockResult], n_days: int):
    ws = wb.create_sheet(sheet_name)

    headers = ["Symbol", "Streak Type", "Net % Change"]
    for i in range(1, n_days + 1):
        headers.append(f"Day {i} % Change")
        headers.append(f"Day {i} Volume")
        headers.append(f"Day {i} Volume % Change")
    headers += ["Next Day % Change", "Next Day Volume", "Verdict"]

    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for stock in stocks:
        row = [stock.symbol, f"{n_days}-Day {stock.direction} Streak", stock.net_pct_change]
        for day in stock.days:
            row.append(day.pct_change)
            row.append(day.volume)
            row.append(day.volume_pct_change if day.volume_pct_change is not None else "N/A")
        row.append(stock.next_day_pct_change if stock.next_day_pct_change is not None else "N/A")
        row.append(stock.next_day_volume if stock.next_day_volume is not None else "N/A")
        row.append(stock.verdict)
        ws.append(row)

        row_fill = green_fill if stock.direction == "Up" else red_fill
        for cell in ws[ws.max_row]:
            cell.fill = row_fill

        if "broken" in stock.verdict.lower():
            verdict_cell = ws.cell(row=ws.max_row, column=len(headers))
            verdict_cell.fill = yellow_fill
            verdict_cell.font = Font(bold=True)

    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    if not stocks:
        ws.cell(row=2, column=1, value="No qualifying stocks for this window.")


def build_excel(report: ReportResponse, n_days: int, end_date: str) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)  # drop default blank sheet

    _write_sheet(wb, f"{n_days}-Day Up Streak", report.up_streaks, n_days)
    _write_sheet(wb, f"{n_days}-Day Down Streak", report.down_streaks, n_days)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ----------------------------------------------------------------------
# API routes
# ----------------------------------------------------------------------

@app.post("/api/report", response_model=ReportResponse)
def get_report(req: ReportRequest):
    return build_report(req)


@app.post("/api/report/stream")
def stream_report(req: ReportRequest):
    """Same computation as /api/report, but streams progress events as
    each symbol finishes checking, then a final 'done' event with the
    full result -- lets the frontend show a real progress bar instead
    of a blind spinner."""

    def event_stream():
        try:
            datetime.strptime(req.end_date, "%Y-%m-%d")
        except ValueError:
            payload = json.dumps({"detail": "end_date must be in YYYY-MM-DD format"})
            yield f"event: error\ndata: {payload}\n\n"
            return

        symbols = req.symbols if req.symbols else get_default_symbols(req.end_date)
        symbols = [s.strip().upper() for s in symbols if s.strip()]
        total = len(symbols)

        if total == 0:
            payload = json.dumps({"detail": "No symbols to check"})
            yield f"event: error\ndata: {payload}\n\n"
            return

        up_streaks, down_streaks = [], []
        skipped = 0
        checked = 0
        # Throttle progress events on very large lists so we don't send
        # thousands of tiny SSE messages; always send the first and last.
        report_every = max(1, total // 200)

        with ThreadPoolExecutor(max_workers=20) as pool:
            futures = {
                pool.submit(evaluate_symbol, sym, req.end_date, req.n_days): sym
                for sym in symbols
            }
            for future in as_completed(futures):
                checked += 1
                try:
                    result = future.result()
                except Exception:
                    result = None

                if result is None:
                    skipped += 1
                elif result.direction == "Up":
                    up_streaks.append(result)
                else:
                    down_streaks.append(result)

                if checked % report_every == 0 or checked == total:
                    payload = json.dumps({"checked": checked, "total": total})
                    yield f"event: progress\ndata: {payload}\n\n"

        up_streaks.sort(key=lambda r: r.net_pct_change, reverse=True)
        down_streaks.sort(key=lambda r: r.net_pct_change)

        report = ReportResponse(
            up_streaks=up_streaks,
            down_streaks=down_streaks,
            checked=total,
            skipped=skipped,
        )
        yield f"event: done\ndata: {report.model_dump_json()}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.post("/api/report/excel")
def get_report_excel(req: ReportRequest):
    report = build_report(req)
    buffer = build_excel(report, req.n_days, req.end_date)
    filename = f"nse_streak_report_{req.end_date}_n{req.n_days}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/default-symbols")
def get_default_symbols_route(end_date: Optional[str] = None):
    symbols = get_default_symbols(end_date)
    return {"symbols": symbols}


_index_list_cache: dict = {}  # name -> {"symbols": [...], "fetched_at": datetime}


@app.get("/api/index-presets")
def get_index_presets():
    return {"presets": list(INDEX_PRESETS.keys())}


@app.get("/api/index-list")
def get_index_list(name: str):
    cached = _index_list_cache.get(name)
    now = datetime.utcnow()
    if cached and (now - cached["fetched_at"]).total_seconds() < 86400:
        return {"symbols": cached["symbols"]}

    try:
        symbols = fetch_index_constituents(name)
    except ValueError:
        raise HTTPException(status_code=404, detail=f"Unknown index preset: {name}")
    except RuntimeError as e:
        raise HTTPException(status_code=502, detail=str(e))

    _index_list_cache[name] = {"symbols": symbols, "fetched_at": now}
    return {"symbols": symbols}


# ----------------------------------------------------------------------
# Static frontend
# ----------------------------------------------------------------------

app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/")
def serve_index():
    return FileResponse("static/index.html")
